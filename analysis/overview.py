import pandas as pd
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# --- File setup ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(BASE_DIR, "../merged.csv")
OUT_PATH = os.path.join(BASE_DIR, "output/overview.xlsx")
os.makedirs(os.path.join(BASE_DIR, "output"), exist_ok=True)

# --- Load data ---
df = pd.read_csv(CSV_PATH)

# --- Clean numeric metrics ---
metrics = [
    "Reach", "Likes", "Comments", "Shares", "Saved",
    "Total Interactions", "Plays", "Views", "Follows"
]
for col in metrics:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

# --- Build summary ---
summary = pd.DataFrame([{
    "Total Posts": len(df),
    "Total Reach": df["Reach"].sum() if "Reach" in df else None,
    "Average Reach": df["Reach"].mean() if "Reach" in df else None,

    "Total Likes": df["Likes"].sum() if "Likes" in df else None,
    "Total Comments": df["Comments"].sum() if "Comments" in df else None,
    "Total Shares": df["Shares"].sum() if "Shares" in df else None,
    "Total Saves": df["Saved"].sum() if "Saved" in df else None,

    "Total Interactions": df["Total Interactions"].sum() if "Total Interactions" in df else None,

    "Total Views": df["Views"].sum() if "Views" in df else None,
    "Total Plays": df["Plays"].sum() if "Plays" in df else None,

    "Total Follows": df["Follows"].sum() if "Follows" in df else None,
    "Average Follows": df["Follows"].mean() if "Follows" in df else None,
}])

# --- Write Excel with styling ---
with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    summary.to_excel(writer, index=False, sheet_name="Summary")
    df.to_excel(writer, index=False, sheet_name="RawData")

    wb = writer.book
    ws = wb["Summary"]

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style numeric cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # Auto-fit column widths
    for column_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 3

print(f"[ok] wrote: {OUT_PATH}")
print(summary.T)